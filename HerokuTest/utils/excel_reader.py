from openpyxl import load_workbook

def get_url_by_page(file_path, sheet_name, page_name):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        excel_page = str(row[0]).strip().lower()
        input_page = page_name.strip().lower()
        if row[0] == page_name:
            url=str(row[1]).strip()
            return url
    raise ValueError("Page not found")